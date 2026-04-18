import os
from datetime import date, timedelta

import pandas as pd
import streamlit as st
from matplotlib import pyplot as plt
import gspread
from google.oauth2.service_account import Credentials
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
@st.cache_data(ttl=60)
def load_price():
    teeg_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTs6jLT1iBie0Fcm28dPQ_x98Pm61yDGxBnHt85bPjyAUw_144eS0HaIEuejDQwYQ/pub?gid=115078867&single=true&output=csv"
    ariston_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQIpFNDSvIXvCQ4-uSvrHyM0QqXpMO83hn2K7b2tCVGJ8hOR9R199Sd2pKwTCRvVQ/pub?gid=1662607201&single=true&output=csv"

    frames = []

    df1 = pd.read_csv(teeg_url)
    df1.columns = df1.columns.str.strip()
    frames.append(df1)

    try:
        df2 = pd.read_csv(ariston_url)
        df2.columns = df2.columns.str.strip()
        frames.append(df2)
    except Exception:
        st.warning("Прайс Ariston пока не загрузился. Проверь ссылку CSV.")

    df_all = pd.concat(frames, ignore_index=True)
    df_all.columns = df_all.columns.str.strip()
    return df_all
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource
def get_gsheet_client():
    creds_dict = dict(st.secrets["gcp_service_account"])
    credentials = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(credentials)

def append_opt_sales_to_gsheet(df: pd.DataFrame):
    gc = get_gsheet_client()

    spreadsheet_name = st.secrets["google_sheets"]["spreadsheet_name"]
    worksheet_name = st.secrets["google_sheets"]["worksheet_name"]

    sh = gc.open(spreadsheet_name)
    ws = sh.worksheet(worksheet_name)

    rows = df.fillna("").values.tolist()
    ws.append_rows(rows, value_input_option="USER_ENTERED")

st.set_page_config(page_title="Финансовая сводка", layout="wide")

# =========================
# СТИЛИ
# =========================
st.markdown("""
<style>

/* ОСНОВА */
.stApp {
    background: #151922;
    color: #f3f4f6;
}

.block-container {
    padding-top: calc(2.8rem + env(safe-area-inset-top));
    padding-bottom: 2rem;
    max-width: 1400px;
}

a.anchor-link {
    display: none !important;
}

h1, h2, h3 {
    color: #f9fafb;
    letter-spacing: 0.2px;
}

/* ЗАГОЛОВКИ */
.main-title {
    font-size: 34px;
    font-weight: 800;
    color: #f9fafb;
    margin-top: 0;
    margin-bottom: 4px;
    line-height: 1.05;
}

.sub-title {
    font-size: 15px;
    color: #8fa3bf;
    margin-bottom: 18px;
}


/* КАРТОЧКИ */
.section-box {
    background: #1d2330;
    border: 1px solid #2f3747;
    border-radius: 18px;
    padding: 14px 16px;
    margin-bottom: 14px;
    box-shadow: 0 6px 20px rgba(0,0,0,0.18);
}

.card {
    background: #1d2330;
    border: 1px solid #2f3747;
    border-radius: 20px;
    padding: 18px 18px;
    margin-bottom: 14px;
    box-shadow: 0 6px 20px rgba(0,0,0,0.18);
}

.card-title {
    font-size: 13px;
    color: #aab2bf;
    margin-bottom: 10px;
}

.card-value {
    font-size: 30px;
    font-weight: 800;
    color: #f8fafc;
    line-height: 1.1;
}

/* ЦВЕТА */
.value-green { color: #34d399; }
.value-red { color: #f87171; }
.value-blue { color: #60a5fa; }

.small-label {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 8px;
    color: #f9fafb;
}

hr {
    border: none;
    height: 1px;
    background: #2f3747;
    margin: 14px 0;
}

/* ОБЫЧНЫЕ КНОПКИ */
.stButton > button {
    background: #1d2330 !important;
    color: #f3f4f6 !important;
    border: 1px solid #2f3747 !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
    padding: 10px 18px !important;
}

.stButton > button:hover {
    border-color: #4b5568 !important;
    color: #ffffff !important;
}

/* 🔥 КНОПКИ СКАЧИВАНИЯ (ИСПРАВЛЕНО) */
.stDownloadButton > button {
    background: #1d2330 !important;
    color: #f3f4f6 !important;
    border: 1px solid #2f3747 !important;
    border-radius: 14px !important;
    font-weight: 600 !important;
    padding: 10px 18px !important;
}

.stDownloadButton > button:hover {
    border-color: #4b5568 !important;
    color: #ffffff !important;
}

.stDownloadButton > button p {
    color: #f3f4f6 !important;
}

/* ПОЛЯ */
div[data-baseweb="input"] > div,
div[data-baseweb="select"] > div,
div[data-testid="stDateInput"] > div,
div[data-testid="stSelectbox"] > div,
div[data-testid="stTextInput"] > div,
div[data-testid="stTextArea"] > div {
    background: #1d2330 !important;
    border: 1px solid #2f3747 !important;
    border-radius: 14px !important;
    color: #f3f4f6 !important;
}

div[data-baseweb="input"] input,
div[data-testid="stDateInput"] input,
div[data-testid="stTextInput"] input,
textarea {
    color: #f3f4f6 !important;
    -webkit-text-fill-color: #f3f4f6 !important;
}

div[data-baseweb="select"] span {
    color: #f3f4f6 !important;
}

/* ВЫПАДАЮЩИЕ СПИСКИ */
div[data-baseweb="popover"] {
    background: #1d2330 !important;
    border-radius: 12px !important;
}

ul[role="listbox"] {
    background: #1d2330 !important;
    color: #f3f4f6 !important;
    border: 1px solid #2f3747 !important;
}

ul[role="listbox"] li {
    color: #f3f4f6 !important;
    background: #1d2330 !important;
}

ul[role="listbox"] li:hover {
    background: #263042 !important;
}

/* EXPANDER */
div[data-testid="stExpander"] details {
    background: #1d2330 !important;
    border: 1px solid #2f3747 !important;
    border-radius: 18px !important;
    overflow: hidden !important;
}

div[data-testid="stExpander"] details summary {
    background: #1d2330 !important;
    color: #f3f4f6 !important;
    padding: 14px 18px !important;
}

div[data-testid="stExpander"] details[open] summary {
    border-bottom: 1px solid #2f3747 !important;
}

/* TABS */
div[data-testid="stTabs"] button {
    color: #cbd5e1 !important;
    font-weight: 700 !important;
}

div[data-testid="stTabs"] button[aria-selected="true"] {
    color: #ffffff !important;
}

/* MOBILE */
@media (max-width: 768px) {
    .block-container {
        padding-top: calc(4.2rem + env(safe-area-inset-top));
    }

    .main-title {
        font-size: 28px;
    }

    .sub-title {
        font-size: 14px;
    }

    .card-value {
        font-size: 26px;
    }
}

/* Убираем курсор в дате */


</style>
""", unsafe_allow_html=True)

# =========================
# URL
# =========================
SALES_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTVCDzAu1DphzNCs2AzlpsjgJyRfzYWEAicdYbqMEFCcjjcxo4WyjVkcKa2-6G2BDyhM2GaBRx23DvO/pub?gid=1240951053&single=true&output=csv"
EXPENSES_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSYEdrQn9FbW5xYzz_UuvUvOUYxbENvC1JnSE4z00YUTvtCxtnP4sU54J-Vs_40kEcuyQLRm-Ae6B_0/pub?gid=1622934317&single=true&output=csv"

ORDERS_FILE = "orders.xlsx"

# =========================
# ВСПОМОГАТЕЛЬНОЕ
# =========================
def format_money(value: float) -> str:
    try:
        return f"{float(value):,.0f}".replace(",", " ")
    except Exception:
        return "0"

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
    )
    return df

def find_column(df: pd.DataFrame, variants: list[str]) -> str | None:
    lower_map = {str(col).strip().lower(): col for col in df.columns}
    for variant in variants:
        found = lower_map.get(variant.lower())
        if found is not None:
            return found
    return None

def parse_mixed_dates(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()

    parsed_dayfirst = pd.to_datetime(s, errors="coerce", dayfirst=True)

    missing = parsed_dayfirst.isna()
    if missing.any():
        parsed_monthfirst = pd.to_datetime(s[missing], errors="coerce", dayfirst=False)
        parsed_dayfirst.loc[missing] = parsed_monthfirst

    parsed_dayfirst = parsed_dayfirst.where(
        (parsed_dayfirst >= pd.Timestamp("2020-01-01")) &
        (parsed_dayfirst <= pd.Timestamp("2030-12-31"))
    )
    return parsed_dayfirst

def parse_float_text(value: str) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if text == "":
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0

def parse_int_text(value: str, default: int = 1) -> int:
    if value is None:
        return default
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if text == "":
        return default
    try:
        parsed = int(float(text))
        return max(1, parsed)
    except Exception:
        return default

@st.cache_data(ttl=60)
def load_data():
    sales_df = pd.read_csv(SALES_URL)
    expenses_df = pd.read_csv(EXPENSES_URL)
    return normalize_columns(sales_df), normalize_columns(expenses_df)

def load_sales_dataframe(data: pd.DataFrame) -> pd.DataFrame:
    df = normalize_columns(data)

    date_col = find_column(df, ["Дата", "дата"])
    channel_col = find_column(df, ["Канал", "канал"])
    name_col = find_column(df, ["Наименование", "наименование"])
    order_col = find_column(df, ["Номер заказа", "номер заказа"])
    cost_col = find_column(df, ["Себестоимость", "себестоимость"])
    rrc_col = find_column(df, ["РРЦ", "ррц"])
    kaspi_col = find_column(df, ["Комиссия Kaspi", "комиссия kaspi"])
    profit_col = find_column(df, ["Чистая прибыль", "чистая прибыль"])
    comment_col = find_column(df, ["Комментарий", "комментарий", "Комментарии", "комментарии"])
    kaspiy_marker_col = find_column(df, ["Каспий", "каспий"])

    if date_col is None:
        st.error("В таблице продаж не найден столбец 'Дата'.")
        st.write("Найденные столбцы:", list(df.columns))
        st.stop()

    if channel_col is None:
        df["Канал"] = ""
        channel_col = "Канал"

    if name_col is None:
        df["Наименование"] = ""
        name_col = "Наименование"

    if order_col is None:
        df["Номер заказа"] = ""
        order_col = "Номер заказа"

    if cost_col is None:
        df["Себестоимость"] = 0
        cost_col = "Себестоимость"

    if rrc_col is None:
        df["РРЦ"] = 0
        rrc_col = "РРЦ"

    if kaspi_col is None:
        df["Комиссия Kaspi"] = 0
        kaspi_col = "Комиссия Kaspi"

    if comment_col is None:
        df["Комментарий"] = ""
        comment_col = "Комментарий"

    if kaspiy_marker_col is None:
        df["Каспий"] = 0
        kaspiy_marker_col = "Каспий"

    df["Дата"] = parse_mixed_dates(df[date_col])
    df["Канал"] = df[channel_col].fillna("").astype(str).str.strip()
    df["Наименование"] = df[name_col].fillna("").astype(str).str.strip()
    df["Номер заказа"] = df[order_col].fillna("").astype(str).str.strip()
    df["Себестоимость"] = pd.to_numeric(df[cost_col], errors="coerce").fillna(0)
    df["РРЦ"] = pd.to_numeric(df[rrc_col], errors="coerce").fillna(0)
    df["Комиссия Kaspi"] = pd.to_numeric(df[kaspi_col], errors="coerce").fillna(0)

    df["Комментарий"] = df[comment_col].fillna("").astype(str)
    df["Комментарий"] = df["Комментарий"].str.replace("\xa0", "", regex=False)
    df["Комментарий"] = df["Комментарий"].str.strip()

    df["Каспий_маркер"] = pd.to_numeric(df[kaspiy_marker_col], errors="coerce").fillna(0)

    # Автоопределение канала, если колонка пустая
    if df["Канал"].eq("").all():
        kaspi_mask = pd.Series(False, index=df.index)

        if "Комиссия Kaspi" in df.columns:
            kaspi_mask = kaspi_mask | (
                pd.to_numeric(df["Комиссия Kaspi"], errors="coerce").fillna(0) > 0
            )

        if "Номер заказа" in df.columns:
            kaspi_mask = kaspi_mask | (
                df["Номер заказа"].fillna("").astype(str).str.strip() != ""
            )

        if "Каспий_маркер" in df.columns:
            kaspi_mask = kaspi_mask | (
                pd.to_numeric(df["Каспий_маркер"], errors="coerce").fillna(0) > 0
            )

        df.loc[kaspi_mask, "Канал"] = "Каспий"
        df.loc[~kaspi_mask, "Канал"] = "ОПТ"

    if profit_col is not None:
        df["Прибыль"] = pd.to_numeric(df[profit_col], errors="coerce").fillna(0)
    else:
        df["Прибыль"] = df["РРЦ"] - df["Себестоимость"] - df["Комиссия Kaspi"]

    df["Маржа %"] = (
        (df["Прибыль"] / df["РРЦ"] * 100)
        .replace([float("inf"), -float("inf")], 0)
        .fillna(0)
    )

    df["Это Ariston"] = df["Наименование"].str.lower().str.contains("ariston", na=False)
    df["Плюс"] = df["Комментарий"] == "+"
    df["Дата_рус"] = df["Дата"].dt.strftime("%d.%m.%Y")

    return df

def load_expenses_dataframe(data: pd.DataFrame) -> pd.DataFrame:
    exp = normalize_columns(data)

    date_col = find_column(exp, ["Дата", "дата"])
    type_col = find_column(exp, ["Тип расхода", "тип расхода"])
    sum_col = find_column(exp, ["Сумма", "сумма"])

    if date_col is None:
        exp["Дата"] = pd.NaT
    else:
        exp["Дата"] = parse_mixed_dates(exp[date_col])

    if type_col is None:
        exp["Тип расхода"] = ""
    else:
        exp["Тип расхода"] = exp[type_col].fillna("").astype(str).str.strip()

    if sum_col is None:
        exp["Сумма"] = 0
    else:
        exp["Сумма"] = pd.to_numeric(exp[sum_col], errors="coerce").fillna(0)

    exp["Дата_рус"] = exp["Дата"].dt.strftime("%d.%m.%Y")
    return exp

def ensure_orders_file():
    if not os.path.exists(ORDERS_FILE):
        pd.DataFrame(columns=[
            "Дата заказа",
            "Канал продажи",
            "Бренд",
            "Модель",
            "Тип цены",
            "Количество",
            "Цена за шт",
            "Общая сумма",
            "Комментарий",
        ]).to_excel(ORDERS_FILE, index=False)

def load_orders_dataframe() -> pd.DataFrame:
    ensure_orders_file()
    try:
        orders = pd.read_excel(ORDERS_FILE)
        orders.columns = (
            orders.columns.astype(str)
            .str.replace("\ufeff", "", regex=False)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
        )
        if "Дата заказа" in orders.columns:
            orders["Дата заказа"] = pd.to_datetime(orders["Дата заказа"], errors="coerce")
        return orders
    except Exception:
        return pd.DataFrame(columns=[
            "Дата заказа",
            "Канал продажи",
            "Бренд",
            "Модель",
            "Тип цены",
            "Количество",
            "Цена за шт",
            "Общая сумма",
            "Комментарий",
        ])

def save_order_row(row: dict):
    ensure_orders_file()
    orders = load_orders_dataframe()
    updated = pd.concat([orders, pd.DataFrame([row])], ignore_index=True)
    updated.to_excel(ORDERS_FILE, index=False)
def build_invoice_pdf(invoice_df: pd.DataFrame) -> bytes:
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buffer = BytesIO()

    base_dir = os.path.dirname(os.path.abspath(__file__))

    regular_candidates = [
        os.path.join(base_dir, "DejaVuSans.ttf"),
        os.path.join(base_dir, "fonts", "DejaVuSans.ttf"),
    ]
    bold_candidates = [
        os.path.join(base_dir, "DejaVuSans-Bold.ttf"),
        os.path.join(base_dir, "fonts", "DejaVuSans-Bold.ttf"),
    ]

    regular_font_path = next((p for p in regular_candidates if os.path.exists(p)), None)
    bold_font_path = next((p for p in bold_candidates if os.path.exists(p)), None)

    regular_font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"

    if regular_font_path:
        pdfmetrics.registerFont(TTFont("CustomFont", regular_font_path))
        regular_font_name = "CustomFont"

    if bold_font_path:
        pdfmetrics.registerFont(TTFont("CustomFont-Bold", bold_font_path))
        bold_font_name = "CustomFont-Bold"
    elif regular_font_path:
        bold_font_name = "CustomFont"

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontName=bold_font_name,
        fontSize=22,
        leading=26,
        alignment=1,
        textColor=colors.black,
        spaceAfter=8,
    )

    sub_style = ParagraphStyle(
        "SubStyle",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=12,
        leading=14,
        alignment=0,
        textColor=colors.black,
        spaceAfter=10,
    )

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=10,
        leading=12,
        textColor=colors.black,
    )

    cell_center = ParagraphStyle(
        "CellCenter",
        parent=cell_style,
        alignment=1,
    )

    cell_right = ParagraphStyle(
        "CellRight",
        parent=cell_style,
        alignment=2,
    )

    df = invoice_df.copy()

    needed_cols = ["Дата", "Бренд", "Модель", "Количество", "Цена", "Сумма", "Комментарий"]
    for col in needed_cols:
        if col not in df.columns:
            df[col] = ""

    df = df[needed_cols].copy()
    df["Дата"] = df["Дата"].astype(str)
    df["Бренд"] = df["Бренд"].astype(str)
    df["Модель"] = df["Модель"].astype(str)
    df["Комментарий"] = df["Комментарий"].astype(str)
    df["Количество"] = pd.to_numeric(df["Количество"], errors="coerce").fillna(0).astype(int)
    df["Цена"] = pd.to_numeric(df["Цена"], errors="coerce").fillna(0)
    df["Сумма"] = pd.to_numeric(df["Сумма"], errors="coerce").fillna(0)

    total_sum = df["Сумма"].sum()

    data = [[
        Paragraph("<b>Дата</b>", cell_style),
        Paragraph("<b>Бренд</b>", cell_style),
        Paragraph("<b>Модель</b>", cell_style),
        Paragraph("<b>Кол-во</b>", cell_center),
        Paragraph("<b>Цена</b>", cell_right),
        Paragraph("<b>Сумма</b>", cell_right),
        Paragraph("<b>Комментарий</b>", cell_style),
    ]]

    for _, row in df.iterrows():
        data.append([
            Paragraph(str(row["Дата"]), cell_style),
            Paragraph(str(row["Бренд"]), cell_style),
            Paragraph(str(row["Модель"]), cell_style),
            Paragraph(str(row["Количество"]), cell_center),
            Paragraph(f"{int(row['Цена']):,}".replace(",", " "), cell_right),
            Paragraph(f"{int(row['Сумма']):,}".replace(",", " "), cell_right),
            Paragraph(str(row["Комментарий"]), cell_style),
        ])

    data.append([
        Paragraph("<b>ИТОГО</b>", cell_style),
        "",
        "",
        "",
        "",
        Paragraph(f"<b>{int(total_sum):,}</b>".replace(",", " "), cell_right),
        "",
    ])

    col_widths = [32 * mm, 28 * mm, 78 * mm, 20 * mm, 28 * mm, 30 * mm, 42 * mm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F79C7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEADING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#8A8A8A")),
        ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DCEAF7")),
        ("SPAN", (0, -1), (4, -1)),
        ("ALIGN", (3, 1), (3, -2), "CENTER"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements = [
        Paragraph("Королевство бойлеров", title_style),
        Paragraph(f"Накладная от {pd.Timestamp.today().strftime('%d.%m.%Y')}", sub_style),
        Spacer(1, 4),
        table,
    ]

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf



    # =========================
# ЗАГРУЗКА
# =========================
sales_raw, expenses_raw = load_data()

base_df = load_sales_dataframe(sales_raw)
base_exp = load_expenses_dataframe(expenses_raw)

if base_df is None or base_df.empty:
    st.error("Не удалось загрузить продажи.")
    st.stop()

if "Дата" not in base_df.columns:
    st.error("В таблице продаж нет колонки 'Дата'.")
    st.write(base_df.columns.tolist())
    st.stop()

valid_dates = base_df["Дата"].dropna()

if valid_dates.empty:
    st.error("В продажах не распознаны даты.")
    st.stop()

# =========================
# ВКЛАДКИ
# =========================
tab1, tab2 = st.tabs(["Финансовая сводка", "Создать заказ"])

# =========================
# ФИНАНСОВАЯ СВОДКА
# =========================
with tab1:
    df = base_df.copy()
    exp = base_exp.copy()

    min_date = valid_dates.min().date()
    safe_today = date.today()

    channel_values = sorted([
        str(x).strip()
        for x in df["Канал"].dropna().unique().tolist()
        if str(x).strip() != ""
    ])
    channel_options = ["Все"] + channel_values

    if "date_from_filter" not in st.session_state:
        st.session_state["date_from_filter"] = min_date

    if "date_to_filter" not in st.session_state:
        st.session_state["date_to_filter"] = safe_today

    # =========================
    # ШАПКА
    # =========================
    left_head, right_head = st.columns([3, 1])

    with left_head:
        st.markdown('<div class="main-title">Панель управления</div>', unsafe_allow_html=True)
        st.markdown('<div class="sub-title">Продажи • Доход • Заказы</div>', unsafe_allow_html=True)

    with right_head:
        st.caption("Кэш ~ 60 сек")
        if st.button("Обновить", use_container_width=True, key="refresh_report"):
            st.cache_data.clear()
            st.rerun()

    search_text = st.text_input("🔍 Поиск по товарам", key="home_search")

    # =========================
    # БЫСТРЫЕ ДЕЙСТВИЯ
    # =========================
    st.markdown("### Быстрые действия")

    a1, a2, a3 = st.columns(3)
    with a1:
        st.markdown("""
        <div class="card">
            <div class="card-title">📦 Остатки</div>
            <div class="card-value" style="font-size:22px;">Склад</div>
        </div>
        """, unsafe_allow_html=True)

    with a2:
        st.markdown("""
        <div class="card">
            <div class="card-title">🧾 Заказы</div>
            <div class="card-value" style="font-size:22px;">Накладные</div>
        </div>
        """, unsafe_allow_html=True)

    with a3:
        st.markdown("""
        <div class="card">
            <div class="card-title">📊 Аналитика</div>
            <div class="card-value" style="font-size:22px;">Сводка</div>
        </div>
        """, unsafe_allow_html=True)

    a4, a5, a6 = st.columns(3)
    with a4:
        st.markdown("""
        <div class="card">
            <div class="card-title">🚚 Продажи</div>
            <div class="card-value" style="font-size:22px;">Движение</div>
        </div>
        """, unsafe_allow_html=True)

    with a5:
        st.markdown("""
        <div class="card">
            <div class="card-title">💰 Доход</div>
            <div class="card-value" style="font-size:22px;">Финансы</div>
        </div>
        """, unsafe_allow_html=True)

    with a6:
        st.markdown("""
        <div class="card">
            <div class="card-title">⚙️ Настройки</div>
            <div class="card-value" style="font-size:22px;">Параметры</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # =========================
    # ФИЛЬТРЫ
    # =========================
    st.markdown("### Фильтры")

    f1, f2 = st.columns(2)

    with f1:
        selected_channel = st.selectbox(
            "Канал",
            channel_options,
            index=0,
            key="report_channel"
        )

    with f2:
        st.markdown("&nbsp;", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        if st.button("Сегодня", use_container_width=True, key="period_today"):
            st.session_state["date_from_filter"] = safe_today
            st.session_state["date_to_filter"] = safe_today
            st.rerun()

    with c2:
        if st.button("7 дней", use_container_width=True, key="period_7"):
            end_date = safe_today
            start_date = max(min_date, end_date - timedelta(days=6))
            st.session_state["date_from_filter"] = start_date
            st.session_state["date_to_filter"] = end_date
            st.rerun()

    with c3:
        if st.button("30 дней", use_container_width=True, key="period_30"):
            end_date = safe_today
            start_date = max(min_date, end_date - timedelta(days=29))
            st.session_state["date_from_filter"] = start_date
            st.session_state["date_to_filter"] = end_date
            st.rerun()

    with c4:
        if st.button("Всё", use_container_width=True, key="period_all"):
            st.session_state["date_from_filter"] = min_date
            st.session_state["date_to_filter"] = safe_today
            st.rerun()

    d1, d2 = st.columns(2)

    with d1:
        date_from = st.date_input(
            "С",
            key="date_from_filter",
            min_value=min_date,
            max_value=safe_today,
            format="YYYY/MM/DD"
        )

    with d2:
        date_to = st.date_input(
            "По",
            key="date_to_filter",
            min_value=min_date,
            max_value=safe_today,
            format="YYYY/MM/DD"
        )

    # =========================
    # ПРИМЕНЕНИЕ ФИЛЬТРОВ
    # =========================
    df = df[
        (df["Дата"].dt.date >= date_from) &
        (df["Дата"].dt.date <= date_to)
    ].copy()

    if selected_channel != "Все":
        df = df[df["Канал"].astype(str).str.strip() == selected_channel].copy()

    exp = exp[
        (exp["Дата"].dt.date >= date_from) &
        (exp["Дата"].dt.date <= date_to)
    ].copy()

    # Поиск по товару
    if search_text.strip():
        search_lower = search_text.strip().lower()
        df = df[
            df["Наименование"].astype(str).str.lower().str.contains(search_lower, na=False)
        ].copy()

    # =========================
    # РАСЧЕТЫ
    # =========================
    df["Мой"] = 0.0
    df.loc[df["Это Ariston"], "Мой"] = df.loc[df["Это Ariston"], "Прибыль"] / 2
    df.loc[~df["Это Ariston"] & df["Плюс"], "Мой"] = df.loc[~df["Это Ariston"] & df["Плюс"], "Прибыль"] / 2

    df["Алексей"] = 0.0
    df.loc[df["Это Ariston"], "Алексей"] = df.loc[df["Это Ariston"], "Прибыль"] / 2
    df.loc[~df["Это Ariston"] & df["Плюс"], "Алексей"] = df.loc[~df["Это Ariston"] & df["Плюс"], "Прибыль"] / 2
    df.loc[~df["Это Ariston"] & ~df["Плюс"], "Алексей"] = df.loc[~df["Это Ariston"] & ~df["Плюс"], "Прибыль"]

    gross_profit = df["Прибыль"].sum()
    my_income = df["Мой"].sum()
    alex_income = df["Алексей"].sum()

    expenses = exp["Сумма"].sum() if "Сумма" in exp.columns else 0
    half_expenses = expenses / 2

    my_net = my_income - half_expenses
    alex_net = alex_income - half_expenses
    total_net = my_net + alex_net

    sales_count = len(df)
    avg_check = df["РРЦ"].mean() if sales_count > 0 else 0
    revenue_sum = df["РРЦ"].sum() if "РРЦ" in df.columns else 0
    margin_percent = (gross_profit / revenue_sum * 100) if revenue_sum > 0 else 0

    kaspi_count = len(df[df["Канал"].astype(str).str.strip() == "Каспий"])
    opt_count = len(df[df["Канал"].astype(str).str.strip() == "ОПТ"])

    # =========================
    # СВОДКА КАК В МАКЕТЕ
    # =========================
    st.markdown("### Сводка")

    s1, s2, s3, s4 = st.columns(4)

    with s1:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Чистая прибыль</div>
            <div class="card-value value-green">{format_money(total_net)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    with s2:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Продаж</div>
            <div class="card-value">{sales_count}</div>
        </div>
        """, unsafe_allow_html=True)

    with s3:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Каспий</div>
            <div class="card-value value-blue">{kaspi_count}</div>
        </div>
        """, unsafe_allow_html=True)

    with s4:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">ОПТ</div>
            <div class="card-value">{opt_count}</div>
        </div>
        """, unsafe_allow_html=True)

    s5, s6, s7 = st.columns(3)

    with s5:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Стас чистый доход</div>
            <div class="card-value">{format_money(my_net)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    with s6:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Алексей чистый доход</div>
            <div class="card-value value-blue">{format_money(alex_net)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    with s7:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Средний чек</div>
            <div class="card-value">{format_money(avg_check)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    # =========================
    # ГРАФИК
    # =========================
    st.markdown("### Динамика прибыли")

    if not df.empty:
        daily_df = (
            df.groupby("Дата", as_index=False)["Прибыль"]
            .sum()
            .sort_values("Дата")
        )

        if not daily_df.empty:
            labels = daily_df["Дата"].dt.strftime("%d.%m")

            fig, ax = plt.subplots(figsize=(10, 4))
            fig.patch.set_facecolor("#151922")
            ax.set_facecolor("#1d2330")

            ax.plot(
                daily_df["Дата"],
                daily_df["Прибыль"],
                marker="o",
                color="#34d399",
                linewidth=2.5
            )

            ax.set_xlabel("")
            ax.set_ylabel("")
            ax.tick_params(colors="#cbd5e1")
            ax.grid(True, alpha=0.18, color="#2f3747")

            for spine in ax.spines.values():
                spine.set_color("#2f3747")

            ax.set_xticks(daily_df["Дата"])
            ax.set_xticklabels(labels, rotation=45, ha="right")

            plt.tight_layout()
            st.pyplot(fig, use_container_width=True)
        else:
            st.info("Нет данных для графика.")
    else:
        st.info("Нет данных для графика.")

    # =========================
    # ТОП ТОВАРОВ
    # =========================
    st.markdown("### Топ-5 товаров")

    if not df.empty:
        top_df = (
            df.groupby("Наименование", as_index=False)["Прибыль"]
            .sum()
            .sort_values("Прибыль", ascending=False)
            .head(5)
        )

        if not top_df.empty:
            for _, row in top_df.iterrows():
                st.markdown(f"""
                <div class="section-box">
                    <div style="font-size:16px; font-weight:700; color:#f3f4f6; margin-bottom:6px;">
                        {row["Наименование"]}
                    </div>
                    <div style="font-size:14px; color:#aab2bf;">
                        Прибыль: <span style="color:#60a5fa; font-weight:700;">{format_money(row["Прибыль"])} ₸</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Нет данных по товарам.")
    else:
        st.info("Нет данных по товарам.")

    # =========================
    # ДОП. БЛОКИ
    # =========================
    e1, e2, e3 = st.columns(3)

    start_date_text = date_from.strftime("%d.%m.%Y")
    end_date_text = date_to.strftime("%d.%m.%Y")

    with e1:
        with st.expander("Быстрый отчет"):
            st.markdown(
                f"""
                <div class="section-box">
                    <div style="font-size:14px; color:#aab2bf; margin-bottom:10px;">
                        Период: <span style="color:#34d399;">{start_date_text} — {end_date_text}</span>
                    </div>
                    <div style="font-size:14px; color:#aab2bf; margin-bottom:12px;">
                        Канал: <span style="color:#f3f4f6;">{selected_channel}</span>
                    </div>
                    <div style="display:flex; justify-content:space-between; margin-bottom:10px;">
                        <span style="color:#aab2bf;">Итого</span>
                        <span style="color:#34d399; font-weight:700;">{format_money(total_net)} ₸</span>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    with e2:
        with st.expander("Продажи"):
            st.markdown(f"""
            <div class="section-box">
                <div style="display:flex; justify-content:space-between; margin-bottom:12px;">
                    <span style="color:#aab2bf;">Количество продаж</span>
                    <span style="font-weight:700;">{sales_count}</span>
                </div>
                <div style="display:flex; justify-content:space-between; margin-bottom:12px;">
                    <span style="color:#aab2bf;">Средний чек</span>
                    <span style="font-weight:700; color:#34d399;">{format_money(avg_check)} ₸</span>
                </div>
                <div style="display:flex; justify-content:space-between;">
                    <span style="color:#aab2bf;">Средняя маржа</span>
                    <span style="font-weight:700; color:#60a5fa;">{margin_percent:.1f}%</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

    with e3:
        with st.expander("Расходы"):
            st.markdown(f"""
            <div class="section-box">
                <div style="font-size:14px; color:#aab2bf;">Общие расходы</div>
                <div style="font-size:28px; font-weight:700; color:#f87171;">{format_money(expenses)} ₸</div>
            </div>
            """, unsafe_allow_html=True)

            if not exp.empty and {"Дата_рус", "Тип расхода", "Сумма"}.issubset(exp.columns):
                recent_exp = exp[["Дата_рус", "Тип расхода", "Сумма"]].tail(3).copy()

                st.markdown("**Последние расходы**")

                for _, row in recent_exp.iterrows():
                    st.markdown(f"""
                    <div class="section-box">
                        <div style="font-size:13px; color:#aab2bf;">{row["Дата_рус"]}</div>
                        <div style="font-size:15px; color:#f3f4f6;">{row["Тип расхода"]}</div>
                        <div style="font-size:16px; font-weight:700; color:#f87171;">{format_money(row["Сумма"])} ₸</div>
                    </div>
                    """, unsafe_allow_html=True)


# =========================
# СОЗДАНИЕ ЗАКАЗА
# =========================
with tab2:
    if "invoice_items" not in st.session_state:
        st.session_state.invoice_items = []

    if "saved_invoice_ready" not in st.session_state:
        st.session_state.saved_invoice_ready = False

    if "invoice_pdf_bytes" not in st.session_state:
        st.session_state.invoice_pdf_bytes = None

    st.markdown('<div class="main-title">Создать заказ</div>', unsafe_allow_html=True)

    price_df = load_price().fillna("").copy()

    for col in ["Бренд", "Модель", "ТипЦены"]:
        if col in price_df.columns:
            price_df[col] = (
                price_df[col]
                .astype(str)
                .str.replace("\xa0", " ", regex=False)
                .str.replace("\ufeff", "", regex=False)
                .str.strip()
            )

    if "Цена" in price_df.columns:
        price_df["Цена"] = pd.to_numeric(price_df["Цена"], errors="coerce").fillna(0)
    else:
        price_df["Цена"] = 0

    if "Себестоимость" in price_df.columns:
        price_df["Себестоимость"] = pd.to_numeric(price_df["Себестоимость"], errors="coerce").fillna(0)
    else:
        price_df["Себестоимость"] = 0

    brands = sorted([
        x for x in price_df["Бренд"].dropna().unique()
        if str(x).strip() != ""
    ])

    brand = st.selectbox("Бренд", brands, key="order_brand")

    models = sorted(
        set(
            price_df.loc[
                price_df["Бренд"].astype(str).str.strip() == str(brand).strip(),
                "Модель"
            ]
            .dropna()
            .astype(str)
            .str.strip()
            .tolist()
        )
    )

    search = st.text_input("🔍 Поиск модели", key="order_model_search")

    if search.strip():
        search_parts = search.lower().strip().split()
        filtered_models = [
            m for m in models
            if all(part in m.lower() for part in search_parts)
        ]
    else:
        filtered_models = models

    if not models:
        st.warning("Для этого бренда нет моделей")
        model = None
    else:
        if search.strip() and not filtered_models:
            st.warning("Модель не найдена. Ниже показан полный список моделей бренда.")
            model_options = models
        else:
            model_options = filtered_models

        model = st.selectbox("Модель", model_options, key="order_model")

    if model:
        price_types = sorted(
            set(
                price_df.loc[
                    (price_df["Бренд"].astype(str).str.strip() == str(brand).strip()) &
                    (price_df["Модель"].astype(str).str.strip() == str(model).strip()),
                    "ТипЦены"
                ]
                .dropna()
                .astype(str)
                .str.strip()
                .tolist()
            )
        )
    else:
        price_types = []

    if price_types:
        price_type = st.selectbox("Тип цены", price_types, key="order_price_type")
    else:
        price_type = None
        st.warning("Для этой модели не найден тип цены")

    if model and price_type:
        selected_row = price_df[
            (price_df["Бренд"].astype(str).str.strip() == str(brand).strip()) &
            (price_df["Модель"].astype(str).str.strip() == str(model).strip()) &
            (price_df["ТипЦены"].astype(str).str.strip() == str(price_type).strip())
        ].copy()
    else:
        selected_row = pd.DataFrame()

    if not selected_row.empty:
        selected_row = selected_row[selected_row["Цена"] > 0]

    price = float(selected_row["Цена"].iloc[0]) if not selected_row.empty else 0
    cost = float(selected_row["Себестоимость"].iloc[0]) if not selected_row.empty else 0

    qty = st.number_input("Количество", min_value=1, value=1, step=1, key="order_qty")
    total_sum = price * qty if price else 0

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Цена</div>
            <div class="card-value value-blue">{format_money(price)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Себестоимость</div>
            <div class="card-value">{format_money(cost)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Сумма</div>
            <div class="card-value">{format_money(total_sum)} ₸</div>
        </div>
        """, unsafe_allow_html=True)

    comment = st.text_input("Комментарий", value="", key="order_comment")

    current_row = {
        "Дата": pd.Timestamp.today().strftime("%d.%m.%Y"),
        "Бренд": brand,
        "Модель": model if model else "",
        "Количество": int(qty),
        "Цена": float(price),
        "Сумма": float(total_sum),
        "Себестоимость": float(cost),
        "Комментарий": comment,
    }

    b1, b2, b3 = st.columns(3)

    with b1:
        if st.button("Добавить позицию", use_container_width=True, key="add_invoice_row"):
            if not model:
                st.warning("Сначала выбери модель")
            elif not price_type:
                st.warning("Сначала выбери тип цены")
            elif price <= 0:
                st.warning("Для выбранной позиции не найдена цена")
            else:
                st.session_state.invoice_items.append(current_row.copy())
                st.session_state.saved_invoice_ready = False
                st.session_state.invoice_pdf_bytes = None
                st.rerun()

    with b2:
        if st.button("Очистить накладную", use_container_width=True, key="clear_invoice"):
            st.session_state.invoice_items = []
            st.session_state.saved_invoice_ready = False
            st.session_state.invoice_pdf_bytes = None
            st.rerun()

    with b3:
        if st.button("Сохранить накладную", use_container_width=True, key="save_invoice"):
            if st.session_state.invoice_items:
                file_path = "orders.xlsx"
                invoice_df = pd.DataFrame(st.session_state.invoice_items).copy()

                final_columns = [
                    "Дата",
                    "Бренд",
                    "Модель",
                    "Количество",
                    "Цена",
                    "Сумма",
                    "Комментарий",
                ]

                for col in final_columns:
                    if col not in invoice_df.columns:
                        invoice_df[col] = ""

                invoice_df = invoice_df[final_columns].copy()
                invoice_df["Количество"] = pd.to_numeric(invoice_df["Количество"], errors="coerce").fillna(1).astype(int)
                invoice_df["Цена"] = pd.to_numeric(invoice_df["Цена"], errors="coerce").fillna(0)
                invoice_df["Сумма"] = pd.to_numeric(invoice_df["Сумма"], errors="coerce").fillna(0)

                total_invoice_sum = invoice_df["Сумма"].sum()

                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Накладная"

                ws.merge_cells("A1:G1")
                ws["A1"] = "Королевство бойлеров"
                ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
                ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

                ws.merge_cells("A2:G2")
                ws["A2"] = f"Накладная от {pd.Timestamp.today().strftime('%d.%m.%Y')}"
                ws["A2"].font = Font(size=11, bold=True, color="FFFFFF")
                ws["A2"].fill = PatternFill("solid", fgColor="4F81BD")
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

                headers = ["Дата", "Бренд", "Модель", "Количество", "Цена", "Сумма", "Комментарий"]
                header_row = 4

                thin = Side(style="thin", color="BFBFBF")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                start_row = 5
                for row_idx, row in enumerate(invoice_df.itertuples(index=False), start_row):
                    values = list(row)
                    for col_num, value in enumerate(values, 1):
                        cell = ws.cell(row=row_idx, column=col_num, value=value)
                        cell.border = border
                        if col_num == 4:
                            cell.alignment = Alignment(horizontal="center")
                        elif col_num in [5, 6]:
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left")

                total_row = start_row + len(invoice_df)

                ws.cell(row=total_row, column=1, value="ИТОГО")
                ws.cell(row=total_row, column=6, value=total_invoice_sum)

                for col_num in range(1, 8):
                    cell = ws.cell(row=total_row, column=col_num)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                    cell.border = border

                ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")

                widths = {
                    "A": 14,
                    "B": 16,
                    "C": 38,
                    "D": 14,
                    "E": 14,
                    "F": 16,
                    "G": 22,
                }

                for col_letter, width in widths.items():
                    ws.column_dimensions[col_letter].width = width

                ws.row_dimensions[1].height = 24
                ws.row_dimensions[2].height = 20

                wb.save(file_path)

                st.session_state.invoice_pdf_bytes = build_invoice_pdf(invoice_df)
                st.session_state.saved_invoice_ready = True
                st.success("Накладная сохранена")
            else:
                st.warning("Накладная пустая")

    st.markdown("---")
    st.markdown("### Позиции в накладной")

    if st.session_state.invoice_items:
        total_invoice_sum = 0.0

        for i in range(len(st.session_state.invoice_items)):
            item = st.session_state.invoice_items[i]

            current_qty = parse_int_text(item.get("Количество", 1), default=1)
            current_price = parse_float_text(item.get("Цена", 0))
            current_sum = current_price * current_qty
            st.session_state.invoice_items[i]["Количество"] = current_qty
            st.session_state.invoice_items[i]["Сумма"] = current_sum

            box1, box2, box3 = st.columns([6, 2, 1])

            with box1:
                st.markdown(f"""
                <div class="section-box" style="margin-bottom:10px;">
                    <div style="font-size:18px; font-weight:700; color:#f3f4f6; margin-bottom:6px;">
                        {item.get("Бренд", "")} — {item.get("Модель", "")}
                    </div>
                    <div style="font-size:15px; color:#cbd5e1;">
                        Цена: <b>{format_money(current_price)} ₸</b>
                    </div>
                    <div style="font-size:15px; color:#cbd5e1;">
                        Сумма: <b>{format_money(current_sum)} ₸</b>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            with box2:
                new_qty = st.number_input(
                    "Кол-во",
                    min_value=1,
                    value=current_qty,
                    step=1,
                    key=f"invoice_qty_{i}"
                )
                if new_qty != current_qty:
                    st.session_state.invoice_items[i]["Количество"] = int(new_qty)
                    st.session_state.invoice_items[i]["Сумма"] = float(current_price) * int(new_qty)
                    st.session_state.saved_invoice_ready = False
                    st.session_state.invoice_pdf_bytes = None
                    st.rerun()

            with box3:
                st.write("")
                st.write("")
                if st.button("🗑️", key=f"delete_invoice_item_{i}", use_container_width=True):
                    st.session_state.invoice_items.pop(i)
                    st.session_state.saved_invoice_ready = False
                    st.session_state.invoice_pdf_bytes = None
                    st.rerun()

            total_invoice_sum += st.session_state.invoice_items[i]["Сумма"]

        st.markdown(f"""
        <div class="card">
            <div class="card-title">Итого по накладной</div>
            <div class="card-value">{format_money(total_invoice_sum)} ₸</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("Пока нет добавленных позиций")

    if st.session_state.saved_invoice_ready:
        d1, d2 = st.columns(2)

        with d1:
            with open("orders.xlsx", "rb") as f:
                st.download_button(
                    "Скачать Excel",
                    data=f,
                    file_name="orders.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_invoice_excel"
                )

        with d2:
            if st.session_state.invoice_pdf_bytes is not None:
                st.download_button(
                    "Скачать PDF",
                    data=st.session_state.invoice_pdf_bytes,
                    file_name="orders.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="download_invoice_pdf"
                )

    if st.button("+ Добавить в продажи (ОПТ)", use_container_width=True, key="add_invoice_to_sales"):
        if not st.session_state.invoice_items:
            st.warning("Накладная пустая")
        else:
            df_to_save = pd.DataFrame(st.session_state.invoice_items).copy()

            df_to_save["Количество"] = pd.to_numeric(
                df_to_save["Количество"], errors="coerce"
            ).fillna(1).astype(int)

            df_to_save["Цена"] = pd.to_numeric(df_to_save["Цена"], errors="coerce").fillna(0)
            df_to_save["Себестоимость"] = pd.to_numeric(df_to_save["Себестоимость"], errors="coerce").fillna(0)

            df_to_save["Дата"] = pd.to_datetime("today").strftime("%d.%m.%Y")
            df_to_save["Канал"] = "ОПТ"

            df_to_save = df_to_save.rename(columns={
                "Модель": "Наименование",
                "Цена": "РРЦ",
            })

            if "Номер заказа" not in df_to_save.columns:
                df_to_save["Номер заказа"] = ""

            if "Комментарий" not in df_to_save.columns:
                df_to_save["Комментарий"] = ""

            df_to_save["Комиссия Kaspi"] = 0
            df_to_save["РРЦ"] = df_to_save["РРЦ"] * df_to_save["Количество"]
            df_to_save["Себестоимость"] = df_to_save["Себестоимость"] * df_to_save["Количество"]

            df_to_save["Чистая прибыль"] = (
                df_to_save["РРЦ"] - df_to_save["Себестоимость"] - df_to_save["Комиссия Kaspi"]
            )

            save_columns = [
                "Дата",
                "Канал",
                "Наименование",
                "Номер заказа",
                "Себестоимость",
                "РРЦ",
                "Комиссия Kaspi",
                "Чистая прибыль",
                "Комментарий",
            ]

            df_to_save = df_to_save[save_columns].copy()
            df_to_save["Комментарий"] = "'" + df_to_save["Комментарий"].astype(str)

            append_opt_sales_to_gsheet(df_to_save)

            st.success("Продажи добавлены")
            st.session_state.invoice_items = []
            st.session_state.saved_invoice_ready = False
            st.session_state.invoice_pdf_bytes = None
            st.rerun()

